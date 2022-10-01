import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
wb.create_sheet(index=1,title="Demo")

sheet = wb["Demo"]

c1 = sheet.cell(row=1,column=1)
c1.value = "Devlabs"
wb.save("demo.xlsx")