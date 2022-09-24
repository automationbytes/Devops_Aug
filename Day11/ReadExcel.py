import openpyxl

def readExcel(filepath,label,header):
    workbook = openpyxl.load_workbook(filepath)
    print(workbook.sheetnames)

    sheet = workbook["DevopsUniv"]
    # workbook.get_sheet_by_name("DevopsUniv")
    # workbook.get_index(3)
    noofRows = sheet.max_row
    noofCols = sheet.max_column
    print(noofRows)
    print(noofCols)

    Label = []
    Header = []

    for i in range(1, noofRows + 1):
        Label.append(sheet.cell(i, 1).value)
        if Label[i - 1].lower() == label.lower():
            for j in range(1, noofCols + 1):
                Header.append(sheet.cell(1, j).value)
                if Header[j - 1].lower() == header.lower():
                    val = sheet.cell(i, j).value
                    break
            break
    return val

#print(readExcel("Datasheet1.xlsx","Facebook","uRL"))